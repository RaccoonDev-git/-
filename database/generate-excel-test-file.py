#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
生成Excel格式的学生导入测试文件
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment


def create_excel_test_file():
    # 创建工作簿
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "学生数据"

    # 表头
    headers = ["姓名", "学号", "年级", "班级", "专业", "手机号", "备注"]

    # 设置表头样式
    header_fill = PatternFill(
        start_color="4CAF50", end_color="4CAF50", fill_type="solid"
    )
    header_font = Font(bold=True, color="FFFFFF")

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # 测试数据
    test_data = [
        [
            "测试学生EXCEL1",
            "EXCEL001",
            2024,
            "24软工EXCEL1",
            "软件工程",
            "13800011111",
            "Excel格式测试数据1",
        ],
        [
            "测试学生EXCEL2",
            "EXCEL002",
            2024,
            "24软工EXCEL2",
            "软件工程",
            "13800022222",
            "Excel格式测试数据2",
        ],
        [
            "测试学生EXCEL3",
            "EXCEL003",
            2023,
            "23计科EXCEL1",
            "计算机科学",
            "13800033333",
            "Excel格式测试数据3",
        ],
        [
            "测试学生EXCEL4",
            "EXCEL004",
            2023,
            "23计科EXCEL2",
            "计算机科学",
            "13800044444",
            "Excel格式测试数据4",
        ],
        [
            "测试学生EXCEL5",
            "EXCEL005",
            2022,
            "22信息EXCEL1",
            "信息工程",
            "13800055555",
            "Excel格式测试数据5",
        ],
    ]

    # 写入数据
    for row_num, row_data in enumerate(test_data, 2):
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.alignment = Alignment(horizontal="left", vertical="center")

    # 调整列宽
    column_widths = [12, 12, 8, 15, 12, 13, 20]
    for col_num, width in enumerate(column_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = width

    # 保存文件
    filename = "test_student_import.xlsx"
    wb.save(filename)
    print(f"✅ Excel测试文件已生成: {filename}")
    print(f"   包含 {len(test_data)} 条测试学生数据")


if __name__ == "__main__":
    try:
        create_excel_test_file()
    except Exception as e:
        print(f"❌ 生成失败: {e}")
        print("\n请先安装openpyxl库:")
        print("  pip install openpyxl")
