#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
生成用户批量导入测试Excel文件
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import os

def create_sample_excel():
    """创建示例Excel文件"""
    
    # 创建工作簿
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "用户导入模板"
    
    # 设置列宽
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 25
    ws.column_dimensions['E'].width = 15
    
    # 设置标题样式
    header_font = Font(bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # 标题行
    headers = ["用户名", "密码", "角色", "邮箱", "手机"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # 测试数据
    test_data = [
        ["test_student01", "pass123456", "STUDENT", "test_stu01@example.com", "13800001001"],
        ["test_student02", "pass123456", "STUDENT", "test_stu02@example.com", "13800001002"],
        ["test_student03", "pass123456", "STUDENT", "test_stu03@example.com", "13800001003"],
        ["test_student04", "pass123456", "STUDENT", "test_stu04@example.com", "13800001004"],
        ["test_student05", "pass123456", "STUDENT", "test_stu05@example.com", "13800001005"],
        ["test_teacher01", "pass123456", "TEACHER", "test_tea01@example.com", "13800002001"],
        ["test_teacher02", "pass123456", "TEACHER", "test_tea02@example.com", "13800002002"],
        ["test_teacher03", "pass123456", "TEACHER", "test_tea03@example.com", "13800002003"],
        ["test_admin01", "pass123456", "ADMIN", "test_adm01@example.com", "13800003001"],
        ["test_admin02", "pass123456", "ADMIN", "test_adm02@example.com", "13800003002"],
    ]
    
    # 填充数据
    for row_idx, data_row in enumerate(test_data, start=2):
        for col_idx, value in enumerate(data_row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = value
            cell.alignment = Alignment(horizontal="left", vertical="center")
    
    # 保存文件
    output_path = "database/用户批量导入测试.xlsx"
    wb.save(output_path)
    print(f"✅ 测试Excel文件已生成: {output_path}")
    print(f"📊 包含 {len(test_data)} 条测试用户数据")
    print(f"   - 学生(STUDENT): 5人")
    print(f"   - 教师(TEACHER): 3人")
    print(f"   - 管理员(ADMIN): 2人")

def create_error_sample_excel():
    """创建包含错误的示例Excel文件(用于测试错误处理)"""
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "错误测试"
    
    # 标题行
    headers = ["用户名", "密码", "角色", "邮箱", "手机"]
    for col, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=header)
    
    # 包含各种错误的测试数据
    error_data = [
        ["", "pass123", "STUDENT", "test@example.com", "13800000001"],  # 空用户名
        ["error_user1", "", "STUDENT", "test@example.com", "13800000002"],  # 空密码
        ["error_user2", "pass123", "", "test@example.com", "13800000003"],  # 空角色
        ["error_user3", "pass123", "INVALID_ROLE", "test@example.com", "13800000004"],  # 无效角色
        ["normal_user1", "pass123456", "STUDENT", "normal1@example.com", "13800000005"],  # 正常数据
        ["normal_user2", "pass123456", "TEACHER", "normal2@example.com", "13800000006"],  # 正常数据
    ]
    
    for row_idx, data_row in enumerate(error_data, start=2):
        for col_idx, value in enumerate(data_row, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    
    output_path = "database/用户批量导入错误测试.xlsx"
    wb.save(output_path)
    print(f"\n✅ 错误测试Excel文件已生成: {output_path}")
    print(f"📊 包含 {len(error_data)} 条数据(4条错误,2条正常)")

if __name__ == "__main__":
    # 检查openpyxl是否安装
    try:
        import openpyxl
    except ImportError:
        print("❌ 错误: 需要安装openpyxl库")
        print("请运行: pip install openpyxl")
        exit(1)
    
    print("=" * 60)
    print("正在生成Excel测试文件...")
    print("=" * 60)
    
    create_sample_excel()
    create_error_sample_excel()
    
    print("\n" + "=" * 60)
    print("✅ 所有测试文件生成完成!")
    print("=" * 60)
    print("\n使用说明:")
    print("1. 打开 database/用户批量导入测试.xlsx 查看正常导入格式")
    print("2. 在用户管理页面点击'批量导入'按钮")
    print("3. 选择生成的Excel文件进行导入测试")
    print("4. 使用'错误测试.xlsx'测试错误处理功能")
