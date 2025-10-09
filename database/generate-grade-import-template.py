"""
生成包含成绩的Excel导入模板文件
"""

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from datetime import datetime


def create_excel_template():
    # 创建工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = "学生成绩导入模板"

    # 定义表头
    headers = [
        "姓名",
        "学号",
        "年级",
        "班级",
        "专业",
        "手机号",
        "数学",
        "英语",
        "计算机基础",
        "数据结构",
        "备注",
    ]

    # 样式定义
    header_font = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(
        start_color="4472C4", end_color="4472C4", fill_type="solid"
    )
    header_alignment = Alignment(horizontal="center", vertical="center")

    cell_font = Font(name="微软雅黑", size=10)
    cell_alignment = Alignment(horizontal="left", vertical="center")

    border_side = Side(style="thin", color="D0D0D0")
    border = Border(
        left=border_side, right=border_side, top=border_side, bottom=border_side
    )

    # 设置表头
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border

    # 示例数据
    students = [
        [
            "张三",
            "20191001",
            "2019",
            "19软工A1",
            "软件工程",
            "13800138000",
            85,
            90,
            88,
            92,
            "优秀学生",
        ],
        [
            "李四",
            "20191002",
            "2019",
            "19软工A1",
            "软件工程",
            "13900139000",
            78,
            82,
            85,
            80,
            "",
        ],
        [
            "王五",
            "20191003",
            "2019",
            "19软工A2",
            "软件工程",
            "13700137000",
            92,
            88,
            90,
            95,
            "班长",
        ],
        [
            "赵六",
            "20191004",
            "2019",
            "19软工A2",
            "软件工程",
            "13600136000",
            70,
            75,
            72,
            68,
            "",
        ],
        [
            "钱七",
            "20191005",
            "2019",
            "19计科B1",
            "计算机科学",
            "13500135000",
            88,
            85,
            87,
            89,
            "",
        ],
        [
            "孙八",
            "20191006",
            "2019",
            "19计科B1",
            "计算机科学",
            "13400134000",
            82,
            80,
            83,
            85,
            "",
        ],
    ]

    # 填充数据
    for row_num, student in enumerate(students, 2):
        for col_num, value in enumerate(student, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = value
            cell.font = cell_font
            cell.alignment = cell_alignment
            cell.border = border

    # 设置列宽
    column_widths = [12, 12, 8, 15, 15, 15, 10, 10, 15, 12, 20]
    for col_num, width in enumerate(column_widths, 1):
        ws.column_dimensions[chr(64 + col_num)].width = width

    # 设置行高
    ws.row_dimensions[1].height = 25
    for row in range(2, len(students) + 2):
        ws.row_dimensions[row].height = 20

    # 添加说明工作表
    ws_info = wb.create_sheet("使用说明")

    instructions = [
        ["📊 学生成绩导入模板 - 使用说明"],
        [""],
        ["✅ 必填字段"],
        ["  • 姓名：学生姓名"],
        ["  • 学号：唯一学号（不能重复）"],
        [""],
        ["📝 可选字段"],
        ["  • 年级：入学年份，如2019"],
        ["  • 班级：班级名称，如19软工A1"],
        ["  • 专业：专业名称，如软件工程"],
        ["  • 手机号：联系方式"],
        ["  • 备注：其他信息"],
        [""],
        ["📊 成绩字段"],
        ["  • 课程列：数学、英语、计算机基础等列名即为课程名称"],
        ["  • 分数范围：0-150分（默认满分100）"],
        ["  • 支持小数：如85.5分"],
        ["  • 空值处理：空单元格会被跳过"],
        [""],
        ["💡 智能功能"],
        ["  • 自动创建课程：不存在的课程会自动创建"],
        ["  • 自动等级计算：A/B/C/D/F等级"],
        ["  • 重复处理：学号存在时更新成绩"],
        ["  • 课程数量：支持任意多门课程"],
        [""],
        ["🚀 使用步骤"],
        ["  1. 在「学生成绩导入模板」工作表中填写数据"],
        ["  2. 可添加更多课程列（列名即课程名）"],
        ["  3. 保存为Excel文件"],
        ["  4. 在系统中选择「批量导入」上传文件"],
        [""],
        ["⚠️ 注意事项"],
        ["  • 不要修改表头（第一行）"],
        ["  • 学号不能重复"],
        ["  • 分数必须是有效数字"],
        ["  • 课程名称长度2-100字符"],
        [""],
        ["📞 遇到问题？"],
        ["  • 查看导入结果中的错误提示"],
        ["  • 阅读GRADE_IMPORT_GUIDE.md完整文档"],
        ["  • 检查系统日志：backend/logs/application.log"],
        [""],
        [f"文档版本：2.0"],
        [f"生成时间：{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"],
    ]

    # 填充说明
    for row_num, instruction in enumerate(instructions, 1):
        cell = ws_info.cell(row=row_num, column=1)
        cell.value = instruction[0]

        if row_num == 1:
            cell.font = Font(name="微软雅黑", size=14, bold=True, color="4472C4")
        elif (
            "✅" in instruction[0]
            or "📝" in instruction[0]
            or "📊" in instruction[0]
            or "💡" in instruction[0]
            or "🚀" in instruction[0]
            or "⚠️" in instruction[0]
            or "📞" in instruction[0]
        ):
            cell.font = Font(name="微软雅黑", size=11, bold=True, color="2E75B6")
        else:
            cell.font = Font(name="微软雅黑", size=10)

    ws_info.column_dimensions["A"].width = 80

    # 保存文件
    output_file = "database/student_import_with_grades_template.xlsx"
    wb.save(output_file)
    print(f"✅ Excel模板已生成: {output_file}")
    print(f"📊 包含 {len(students)} 个示例学生")
    print(f"📚 包含 4 门课程成绩")
    print(f"📄 包含使用说明工作表")


if __name__ == "__main__":
    try:
        create_excel_template()
    except Exception as e:
        print(f"❌ 生成失败: {e}")
        import traceback

        traceback.print_exc()
